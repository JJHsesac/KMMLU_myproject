#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
KMMLU A.X-4.0-Light Zero-shot 평가기
====================================
팀 과제용 - 간단하고 확실한 버전

LangSmith 연결 (선택):
export LANGCHAIN_API_KEY="ls-..."  # 있으면 자동 연결
"""

import torch
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import random
import argparse
from tqdm import tqdm
from datasets import load_dataset
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig
from datetime import datetime
import json
import warnings
import os

warnings.filterwarnings('ignore')

# LangSmith 자동 감지
LANGSMITH_ENABLED = False
if os.getenv("LANGCHAIN_API_KEY"):
    try:
        from langsmith import Client
        os.environ["LANGCHAIN_TRACING_V2"] = "true"
        if not os.getenv("LANGCHAIN_PROJECT"):
            os.environ["LANGCHAIN_PROJECT"] = "kmmlu-experiments"
        LANGSMITH_ENABLED = True
        print("✅ LangSmith 활성화됨")
    except ImportError:
        print("⚠️ langsmith 미설치 (pip install langsmith)")
else:
    print("ℹ️ LangSmith 비활성화 (LANGCHAIN_API_KEY 없음)")


class KMMLUEvaluator:
    """A.X-4.0-Light Zero-shot 평가기"""
    
    def __init__(
        self,
        model_id: str = "skt/A.X-4.0-Light",
        batch_size: int = 2,
        seed: int = 42,
        output_prefix: str = "kmmlu_ax_zeroshot",
    ):
        self.model_id = model_id
        self.batch_size = batch_size
        self.seed = seed
        self.output_prefix = output_prefix
        self.langsmith = LANGSMITH_ENABLED
        
        # Random seed
        random.seed(seed)
        torch.manual_seed(seed)
        if torch.cuda.is_available():
            torch.cuda.manual_seed_all(seed)
        
        self.tokenizer, self.model = self._load_model()
        self.subsets = self._get_subsets()
        self.categories = self._get_categories()
        
        self.answer_map = {
            "A": 0, "B": 1, "C": 2, "D": 3,
            "1": 0, "2": 1, "3": 2, "4": 3,
        }
    
    def _load_model(self):
        """모델 로드"""
        print(f"\n{'='*60}")
        print(f"🤖 모델: {self.model_id}")
        print(f"⚡ Zero-shot (num_shots=0)")
        print(f"🎲 Seed: {self.seed}")
        print(f"{'='*60}\n")
        
        # 4bit 양자화
        bnb_config = BitsAndBytesConfig(
            load_in_4bit=True,
            bnb_4bit_use_double_quant=True,
            bnb_4bit_quant_type="nf4",
            bnb_4bit_compute_dtype=torch.bfloat16,
        )
        
        tokenizer = AutoTokenizer.from_pretrained(self.model_id)
        tokenizer.padding_side = "left"
        if tokenizer.pad_token is None:
            tokenizer.pad_token = tokenizer.eos_token
        
        model = AutoModelForCausalLM.from_pretrained(
            self.model_id,
            quantization_config=bnb_config,
            device_map="auto",
        )
        model.eval()
        
        print("✅ 모델 로딩 완료\n")
        return tokenizer, model
    
    def _format_prompt(self, ex):
        """프롬프트 생성"""
        text = f"문제: {ex['question']}\n"
        for choice in ["A", "B", "C", "D"]:
            text += f"{choice}. {ex[choice]}\n"
        text += "정답:"
        return text
    
    def _get_answer_idx(self, ex):
        """정답 인덱스"""
        ans = str(ex.get("answer", "")).strip().upper()
        if ans in self.answer_map:
            return self.answer_map[ans]
        elif ans.isdigit() and 1 <= int(ans) <= 4:
            return int(ans) - 1
        return None
    
    def _predict(self, inputs):
        """예측"""
        with torch.no_grad():
            logits = self.model(**inputs).logits[:, -1, :]
        
        choice_ids = [
            self.tokenizer.encode(ch, add_special_tokens=False)[0]
            for ch in ["A", "B", "C", "D"]
        ]
        
        preds = torch.argmax(logits[:, choice_ids], dim=-1)
        return preds.cpu().tolist()
    
    def evaluate(self):
        """평가 실행"""
        results = []
        total_correct, total_count = 0, 0
        start_time = datetime.now()
        
        print("🚀 평가 시작...\n")
        
        for subset in tqdm(self.subsets, desc="📊 진행"):
            try:
                dataset = load_dataset("HAERAE-HUB/KMMLU", subset)
                
                if "test" not in dataset:
                    continue
                
                test_data = list(dataset["test"])
                correct, count = 0, 0
                
                # 배치 처리
                for i in range(0, len(test_data), self.batch_size):
                    batch = test_data[i:i + self.batch_size]
                    prompts = [self._format_prompt(ex) for ex in batch]
                    
                    inputs = self.tokenizer(
                        prompts,
                        return_tensors="pt",
                        padding=True,
                        truncation=True,
                        max_length=2048,
                    ).to(self.model.device)
                    
                    predictions = self._predict(inputs)
                    
                    for ex, pred in zip(batch, predictions):
                        gt = self._get_answer_idx(ex)
                        if gt is not None:
                            count += 1
                            if pred == gt:
                                correct += 1
                
                # 결과 저장
                accuracy = correct / count if count > 0 else 0
                category = self.categories.get(subset, "Other")
                
                results.append({
                    "Subset": subset,
                    "Category": category,
                    "Correct": correct,
                    "Total": count,
                    "Accuracy": accuracy,
                })
                
                total_correct += correct
                total_count += count
                
            except Exception as e:
                print(f"❌ {subset} 실패: {e}")
                continue
        
        # 결과 저장
        end_time = datetime.now()
        elapsed = end_time - start_time
        
        self._save_all(results, total_correct, total_count, elapsed)
        
        print(f"\n{'='*60}")
        print(f"✅ 완료!")
        print(f"⏱️  시간: {elapsed}")
        print(f"📊 정확도: {total_correct/total_count:.4f} ({total_correct}/{total_count})")
        print(f"{'='*60}\n")
    
    def _save_all(self, results, correct, total, elapsed):
        """결과 저장 (CSV + JSON + XLSX)"""
        df = pd.DataFrame(results)
        overall_acc = correct / total if total > 0 else 0
        cat_acc = df.groupby("Category")["Accuracy"].mean()
        
        # 1. CSV
        csv_file = f"{self.output_prefix}.csv"
        df.to_csv(csv_file, index=False, encoding="utf-8-sig")
        print(f"✅ CSV: {csv_file}")
        
        # 2. JSON
        json_data = {
            "model_id": self.model_id,
            "evaluation_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "time_elapsed": str(elapsed),
            "time_elapsed_seconds": round(elapsed.total_seconds(), 2),
            "experiment_config": {
                "seed": self.seed,
                "batch_size": self.batch_size,
                "num_shots": 0,
                "prompting_strategy": "zero_shot",
            },
            "summary": {
                "overall_accuracy": round(overall_acc, 4),
                "correct_answers": correct,
                "total_questions": total,
                "category_accuracy": {k: round(v, 4) for k, v in cat_acc.to_dict().items()},
            },
            "subset_scores": [
                {
                    "subset": row["Subset"],
                    "category": row["Category"],
                    "accuracy": round(row["Accuracy"], 4),
                    "correct": row["Correct"],
                    "total": row["Total"],
                }
                for _, row in df.iterrows()
            ],
        }
        
        json_file = f"{self.output_prefix}.json"
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON: {json_file}")
        
        # 3. XLSX
        self._save_excel(df, overall_acc, cat_acc)
    
    def _save_excel(self, df, overall_acc, cat_acc):
        """XLSX 저장"""
        xlsx_file = f"{self.output_prefix}_comparison.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 시트 1: 상세
        ws1 = wb.create_sheet("4가지분류상세")
        ws1.append(["분야", "과목", "TEST", "Accuracy"])
        for _, row in df.iterrows():
            ws1.append([row["Category"], row["Subset"], row["Total"], round(row["Accuracy"], 4)])
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            for cell in row:
                cell.border = border
                if cell.column == 4:
                    cell.number_format = '0.0000'
        
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        
        # 시트 2: 분야별
        ws2 = wb.create_sheet("분야별_요약")
        ws2.append(["분야", "과목 수", "평균 정확도"])
        cat_count = df.groupby("Category").size()
        for cat in ["STEM", "HUMSS", "Applied Science", "Other"]:
            if cat in cat_acc:
                ws2.append([cat, cat_count[cat], round(cat_acc[cat], 4)])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                cell.border = border
                if cell.column == 3:
                    cell.number_format = '0.0000'
        
        # 시트 3: 통계
        ws3 = wb.create_sheet("전체_통계")
        ws3.append(["항목", "값"])
        stats = [
            ["모델", self.model_id],
            ["평가 방법", "Zero-shot"],
            ["전체 정확도", round(overall_acc, 4)],
            ["과목 수", len(df)],
            ["총 문제", df["Total"].sum()],
            ["정답 수", df["Correct"].sum()],
        ]
        for item, value in stats:
            ws3.append([item, value])
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
            for cell in row:
                cell.border = border
        
        wb.save(xlsx_file)
        print(f"✅ XLSX: {xlsx_file} (3시트)")
    
    def _get_subsets(self):
        """KMMLU 45개 과목"""
        return [
            "Accounting", "Agricultural-Sciences", "Aviation-Engineering-and-Maintenance",
            "Biology", "Chemical-Engineering", "Chemistry", "Civil-Engineering",
            "Computer-Science", "Construction", "Criminal-Law", "Ecology", "Economics",
            "Education", "Electrical-Engineering", "Electronics-Engineering",
            "Energy-Management", "Environmental-Science", "Fashion", "Food-Processing",
            "Gas-Technology-and-Engineering", "Geomatics", "Health", "Industrial-Engineer",
            "Information-Technology", "Interior-Architecture-and-Design", "Law",
            "Machine-Design-and-Manufacturing", "Management", "Maritime-Engineering",
            "Marketing", "Materials-Engineering", "Mechanical-Engineering",
            "Nondestructive-Testing", "Patent", "Political-Science-and-Sociology",
            "Psychology", "Public-Safety", "Railway-and-Automotive-Engineering",
            "Real-Estate", "Refrigerating-Machinery", "Social-Welfare", "Taxation",
            "Telecommunications-and-Wireless-Technology", "Korean-History", "Math",
        ]
    
    def _get_categories(self):
        """분야 매핑"""
        cats = {
            "STEM": [
                "Biology", "Chemical-Engineering", "Chemistry", "Civil-Engineering",
                "Computer-Science", "Ecology", "Electrical-Engineering",
                "Information-Technology", "Materials-Engineering",
                "Mechanical-Engineering", "Math",
            ],
            "HUMSS": [
                "Accounting", "Criminal-Law", "Economics", "Education", "Law",
                "Management", "Political-Science-and-Sociology", "Psychology",
                "Social-Welfare", "Taxation", "Korean-History",
            ],
            "Applied Science": [
                "Aviation-Engineering-and-Maintenance", "Electronics-Engineering",
                "Energy-Management", "Environmental-Science",
                "Gas-Technology-and-Engineering", "Geomatics", "Industrial-Engineer",
                "Machine-Design-and-Manufacturing", "Maritime-Engineering",
                "Nondestructive-Testing", "Railway-and-Automotive-Engineering",
                "Telecommunications-and-Wireless-Technology",
            ],
            "Other": [
                "Agricultural-Sciences", "Construction", "Fashion", "Food-Processing",
                "Health", "Interior-Architecture-and-Design", "Marketing", "Patent",
                "Public-Safety", "Real-Estate", "Refrigerating-Machinery",
            ],
        }
        return {s: cat for cat, subs in cats.items() for s in subs}


def main():
    parser = argparse.ArgumentParser(description="KMMLU A.X Zero-shot 평가")
    parser.add_argument("--model_id", type=str, default="skt/A.X-4.0-Light")
    parser.add_argument("--batch_size", type=int, default=2)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output_prefix", type=str, default="kmmlu_ax_zeroshot")
    
    args = parser.parse_args()
    
    evaluator = KMMLUEvaluator(
        model_id=args.model_id,
        batch_size=args.batch_size,
        seed=args.seed,
        output_prefix=args.output_prefix,
    )
    
    evaluator.evaluate()


if __name__ == "__main__":
    main()
